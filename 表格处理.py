import streamlit as st
import pandas as pd
import openpyxl
import os
import traceback
import re
import io
import tempfile
import base64


class ExcelProcessor:
    def __init__(self):
        self.monthly_account = None
        self.billing_period = None
        self.order_count = None
        self.file_path = None
        self.total_fee = None
        self.total_discount = None
        self.total_payable = None
        self.total_claims = None
        self.overview_amount = None  # 新增：账单总览金额
        self.special_ticket_discount = None  # 新增：特殊单票折扣

    def process_excel(self, file_bytes, file_name):
        """处理Excel文件，提取所需信息"""
        self.file_path = file_name  # 使用文件名代替完整路径

        try:
            # 创建临时文件以使openpyxl能够处理
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_file.write(file_bytes)
                temp_path = temp_file.name

            # 使用openpyxl加载工作簿以处理合并单元格
            wb = openpyxl.load_workbook(temp_path, data_only=True)

            # 1. 获取月结账号 (从账单总览sheet的D6:G6合并单元格)
            try:
                overview_sheet = wb["账单总览"]
            except KeyError:
                # 如果没有找到账单总览，尝试找其他可能的sheet名
                sheet_names = wb.sheetnames
                overview_sheet = None
                for name in sheet_names:
                    if "总览" in name or "概览" in name:
                        overview_sheet = wb[name]
                        break
                
                if overview_sheet is None:
                    # 如果仍然找不到，使用第一个sheet
                    overview_sheet = wb[sheet_names[0]]

            # 查找合并单元格
            for merged_range in overview_sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row

                # 检查J6:L6合并单元格 (D=4, G=7, 行号为6)
                if min_col == 4 and max_col == 7 and min_row == 6 and max_row == 6:
                    # 从合并单元格的左上角获取值
                    self.monthly_account = overview_sheet.cell(row=6, column=4).value
                    break

            # 2. 获取账单周期 (从账单总览sheet的D7:G7合并单元格)
            for merged_range in overview_sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row

                # 检查D7:G7合并单元格 (D=4, G=7, 行号为7)
                if min_col == 4 and max_col == 7 and min_row == 7 and max_row == 7:
                    # 从合并单元格的左上角获取值
                    self.billing_period = overview_sheet.cell(row=7, column=4).value
                    break

            # 3. 获取当月单量 - 修改为统计N列值为"运费"的行数
            try:
                detail_sheet = wb["账单明细"]
            except KeyError:
                # 如果没有找到账单明细，尝试找其他可能的sheet名
                sheet_names = wb.sheetnames
                detail_sheet = None
                for name in sheet_names:
                    if "明细" in name or "详情" in name:
                        detail_sheet = wb[name]
                        break
                
                if detail_sheet is None:
                    # 如果仍然找不到，使用第一个不是总览的sheet
                    for name in sheet_names:
                        if name != overview_sheet.title:
                            detail_sheet = wb[name]
                            break

            # 查找标题行
            header_row = None
            for row in range(1, 10):  # 在前10行中查找标题
                found_header = False
                for col in range(1, 15):  # 在前15列中查找
                    cell_value = detail_sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and "服务" in cell_value:
                        header_row = row
                        found_header = True
                        break
                if found_header:
                    break
            
            # 统计N列中值为"运费"的行数（N列=14）
            n_column = 14
            freight_count = 0
            
            if header_row:
                for row in range(header_row + 1, detail_sheet.max_row + 1):
                    n_value = detail_sheet.cell(row=row, column=n_column).value
                    if n_value and isinstance(n_value, str) and n_value.strip() == "运费":
                        freight_count += 1
            
            if freight_count > 0:
                self.order_count = freight_count
            else:
                # 如果在N列中没有找到"运费"，回退到备选方法
                # 检查是否有其他列中包含"运费"
                freight_col = None
                if header_row:
                    for col in range(1, detail_sheet.max_column + 1):
                        for row in range(header_row + 1, min(header_row + 10, detail_sheet.max_row + 1)):
                            cell_value = detail_sheet.cell(row=row, column=col).value
                            if cell_value and isinstance(cell_value, str) and cell_value.strip() == "运费":
                                freight_col = col
                                break
                        if freight_col:
                            break
                
                if freight_col:
                    # 如果找到了包含"运费"的列，统计该列中的"运费"行数
                    freight_count = 0
                    for row in range(header_row + 1, detail_sheet.max_row + 1):
                        cell_value = detail_sheet.cell(row=row, column=freight_col).value
                        if cell_value and isinstance(cell_value, str) and cell_value.strip() == "运费":
                            freight_count += 1
                    self.order_count = freight_count
                else:
                    # 如果仍然没找到，回退到原来的方法
                    numeric_values = []
                    for row in range(2, detail_sheet.max_row + 1):
                        cell_value = detail_sheet.cell(row=row, column=1).value
                        if cell_value is not None:
                            try:
                                if isinstance(cell_value, str):
                                    numbers = re.findall(r'\d+', cell_value)
                                    if numbers:
                                        numeric_values.append(int(numbers[0]))
                                elif isinstance(cell_value, (int, float)):
                                    numeric_values.append(int(cell_value))
                            except (ValueError, TypeError):
                                continue

                    if numeric_values:
                        self.order_count = max(numeric_values)
                    else:
                        non_empty_rows = 0
                        for row in range(2, detail_sheet.max_row + 1):
                            if detail_sheet.cell(row=row, column=1).value is not None:
                                non_empty_rows += 1
                        self.order_count = non_empty_rows

            # 4. 获取账单明细中的汇总数据
            # 动态查找汇总数据所在行
            self._find_summary_values(detail_sheet)
            
            # 5. 获取账单总览金额 (从账单总览sheet的J17:L17合并单元格)
            self._find_overview_amount(overview_sheet)
            
            # 6. 获取特殊单票折扣
            self._find_special_ticket_discount(detail_sheet)

            # 删除临时文件
            try:
                os.unlink(temp_path)
            except:
                pass

            return {
                "月结账号": self.monthly_account,
                "账单周期": self.billing_period,
                "当月单量": self.order_count,
                "费用(元)": self.total_fee,
                "折扣/促销": self.total_discount,
                "应付金额": self.total_payable,
                "理赔费用合计": self.total_claims,
                "账单总览金额": self.overview_amount,  # 新增字段
                "特殊单票折扣": self.special_ticket_discount,  # 新增字段
                "文件名": file_name
            }

        except Exception as e:
            # 确保清理临时文件
            try:
                if 'temp_path' in locals():
                    os.unlink(temp_path)
            except:
                pass
            error_message = f"处理文件时出错: {str(e)}\n{traceback.format_exc()}"
            raise Exception(error_message)

    def _find_overview_amount(self, overview_sheet):
        """获取账单总览金额（从账单总览sheet的J17:L17合并单元格）"""
        self.overview_amount = None
        
        try:
            # 首先检查J16:L16是否为合并单元格
            for merged_range in overview_sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
                
                # 检查J16:L16合并单元格 (J=10, L=12, 行号为17)
                if min_col == 10 and max_col == 12 and min_row == 17 and max_row == 17:
                    # 从合并单元格的左上角获取值
                    self.overview_amount = overview_sheet.cell(row=17, column=10).value
                    return
            
            # 如果没有找到合并单元格，尝试直接获取J17单元格的值
            self.overview_amount = overview_sheet.cell(row=17, column=10).value
            
            # 如果还是没有找到值，尝试在总览页面寻找"合计"或"总计"附近的金额
            if not self._is_valid_number(self.overview_amount):
                for row in range(15, 25):  # 在15-25行范围内查找
                    for col in range(1, 15):
                        cell_value = overview_sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str) and ("合计" in cell_value or "总计" in cell_value):
                            # 找到合计行，尝试在同一行的后面几列查找金额
                            for right_col in range(col + 1, col + 5):
                                right_value = overview_sheet.cell(row=row, column=right_col).value
                                if self._is_valid_number(right_value):
                                    self.overview_amount = right_value
                                    return
        except Exception as e:
            print(f"查找账单总览金额时出错: {str(e)}")
            traceback.print_exc()

    def _find_special_ticket_discount(self, detail_sheet):
        """查找特殊单票折扣 - 类似理赔费用的查找逻辑"""
        self.special_ticket_discount = None
        max_row = detail_sheet.max_row
        
        try:
            # 首先检查是否存在"特殊单票折扣"相关单元格
            has_special_discount = False
            for row in range(1, max_row + 1):
                for col in range(1, 20):
                    cell_value = detail_sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and "特殊单票折扣" in cell_value:
                        has_special_discount = True
                        break
                if has_special_discount:
                    break
            
            # 如果存在特殊单票折扣相关单元格，则在D列中找到最大的数字
            if has_special_discount:
                d_column_values = []
                for row in range(2, max_row + 1):  # 从第2行开始，跳过表头
                    cell_value = detail_sheet.cell(row=row, column=4).value  # D列 = 4
                    if self._is_valid_number(cell_value):
                        # 如果是字符串，尝试转换为浮点数
                        if isinstance(cell_value, str):
                            try:
                                cleaned_value = re.sub(r'[^\d.-]', '', cell_value)
                                if cleaned_value:  # 确保不是空字符串
                                    d_column_values.append(float(cleaned_value))
                            except ValueError:
                                pass
                        else:
                            d_column_values.append(float(cell_value))
                
                # 从D列值中找到最大的数值作为特殊单票折扣
                if d_column_values:
                    self.special_ticket_discount = max(d_column_values)
                else:
                    self.special_ticket_discount = None  # 没有找到值，特殊单票折扣为空
            else:
                # 如果不存在特殊单票折扣相关单元格，则设置为None
                self.special_ticket_discount = None
                
        except Exception as e:
            print(f"查找特殊单票折扣时出错: {str(e)}")
            traceback.print_exc()

    def _is_valid_number(self, value):
        """检查值是否为有效数字"""
        if value is None:
            return False

        try:
            # 如果是字符串，尝试转换为浮点数
            if isinstance(value, str):
                # 去除可能的货币符号、括号等
                cleaned_value = re.sub(r'[^\d.-]', '', value)
                if cleaned_value:  # 确保不是空字符串
                    float(cleaned_value)
                    return True
                return False
            # 直接检查是否为数字类型
            return isinstance(value, (int, float))
        except (ValueError, TypeError):
            return False

    def _find_summary_values(self, detail_sheet):
        """查找汇总数据 - 通过搜索文本而不依赖于固定位置"""
        # 初始化默认值为None，表示未找到
        self.total_fee = None
        self.total_discount = None
        self.total_payable = None
        self.total_claims = None

        max_row = detail_sheet.max_row

        try:
            # 从表格的最后30行开始向上查找汇总数据
            search_start = max(2, max_row - 50)

            # 先查找包含"合计"或者"总计"的行
            total_rows = []
            for row in range(max_row, search_start, -1):
                found_total = False
                for col in range(1, 20):  # 扫描更多列以确保能找到关键词
                    cell_value = detail_sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and (
                            "合计" in cell_value or "合 计" in cell_value or "总计" in cell_value):
                        total_rows.append(row)
                        found_total = True
                        break
                if found_total:
                    continue  # 继续搜索其他可能的汇总行

            if total_rows:
                # 找到了可能的汇总行，开始在这些行周围寻找关键字段
                for row in total_rows:
                    # 查找标题行 - 通常在数据的前几行
                    header_row = None
                    fee_col = None  # 初始化fee_col变量
                    discount_col = None
                    payable_col = None

                    for r in range(1, 30):  # 查找更多行以找到标题
                        for c in range(1, 20):
                            header_value = detail_sheet.cell(row=r, column=c).value
                            if header_value and isinstance(header_value, str):
                                # 查找与费用(元)、折扣/促销、应付金额相关的标题
                                if "费用" in header_value and ("元" in header_value or "¥" in header_value):
                                    fee_col = c
                                    header_row = r
                                if "折扣" in header_value or "促销" in header_value:
                                    discount_col = c
                                if "应付" in header_value and "金额" in header_value:
                                    payable_col = c

                    # 如果找到了标题行和至少一个相关列，根据标题位置查找对应的汇总值
                    if header_row and (fee_col or discount_col or payable_col):
                        if fee_col:
                            fee_value = detail_sheet.cell(row=row, column=fee_col).value
                            if self._is_valid_number(fee_value) and self.total_fee is None:
                                self.total_fee = fee_value
                        if discount_col:
                            discount_value = detail_sheet.cell(row=row, column=discount_col).value
                            if self._is_valid_number(discount_value) and self.total_discount is None:
                                self.total_discount = discount_value
                        if payable_col:
                            payable_value = detail_sheet.cell(row=row, column=payable_col).value
                            if self._is_valid_number(payable_value) and self.total_payable is None:
                                self.total_payable = payable_value
                    else:
                        # 如果没有找到标题行，则尝试在合计行查找
                        # 寻找合计行的前一行，检查是否有标题文本
                        prev_row = row - 1
                        if prev_row > 0:
                            for c in range(1, 20):
                                col_header = detail_sheet.cell(row=prev_row, column=c).value
                                if col_header and isinstance(col_header, str):
                                    if "费用" in col_header and ("元" in col_header or "¥" in col_header):
                                        fee_col = c
                                    if "折扣" in col_header or "促销" in col_header:
                                        discount_col = c
                                    if "应付" in col_header and "金额" in col_header:
                                        payable_col = c

                            # 根据找到的列获取对应的汇总值
                            if fee_col:
                                fee_value = detail_sheet.cell(row=row, column=fee_col).value
                                if self._is_valid_number(fee_value) and self.total_fee is None:
                                    self.total_fee = fee_value
                            if discount_col:
                                discount_value = detail_sheet.cell(row=row, column=discount_col).value
                                if self._is_valid_number(discount_value) and self.total_discount is None:
                                    self.total_discount = discount_value
                            if payable_col:
                                payable_value = detail_sheet.cell(row=row, column=payable_col).value
                                if self._is_valid_number(payable_value) and self.total_payable is None:
                                    self.total_payable = payable_value

                        # 如果仍然没有找到主要字段，则尝试从合计行向右查找数值
                        if not all([self.total_fee, self.total_payable]):
                            for right_col in range(1, 20):
                                right_value = detail_sheet.cell(row=row, column=right_col).value
                                if self._is_valid_number(right_value):
                                    # 找到了数值，根据位置依次分配
                                    if self.total_fee is None:
                                        self.total_fee = right_value
                                        continue
                                    if self.total_discount is None:
                                        self.total_discount = right_value
                                        continue
                                    if self.total_payable is None:
                                        self.total_payable = right_value
                                        break

            # 查找理赔费用合计 - 首先检查是否存在理赔费用相关单元格
            has_claims = False
            for row in range(1, max_row + 1):
                for col in range(1, 20):
                    cell_value = detail_sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and "理赔" in cell_value:
                        has_claims = True
                        break
                if has_claims:
                    break
            
            # 如果存在理赔费用相关单元格，则在H列中查找最小的负值
            if has_claims:
                h_column_values = []
                for row in range(2, max_row + 1):  # 从第2行开始，跳过表头
                    cell_value = detail_sheet.cell(row=row, column=8).value  # H列 = 8
                    if self._is_valid_number(cell_value):
                        # 如果是字符串，尝试转换为浮点数
                        if isinstance(cell_value, str):
                            try:
                                cleaned_value = re.sub(r'[^\d.-]', '', cell_value)
                                if cleaned_value:  # 确保不是空字符串
                                    h_column_values.append(float(cleaned_value))
                            except ValueError:
                                pass
                        else:
                            h_column_values.append(float(cell_value))

                # 从H列值中找到最小的负值作为理赔费用
                negative_values = [val for val in h_column_values if val < 0]
                if negative_values:
                    self.total_claims = min(negative_values)  # 取最小的负值
                else:
                    self.total_claims = None  # 没有找到负值，理赔费用为空
            else:
                # 如果不存在理赔费用相关单元格，则设置为None
                self.total_claims = None

        except Exception as e:
            # 记录错误但不中断程序
            print(f"查找汇总数据时出错: {str(e)}")
            traceback.print_exc()


def get_table_download_link(df):
    """生成一个下载链接，允许下载DataFrame作为Excel文件"""
    # 将DataFrame转换为Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    excel_data = output.getvalue()
    
    # 使用base64编码
    b64 = base64.b64encode(excel_data).decode()
    
    # 创建下载链接
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="账单数据提取结果.xlsx">下载Excel文件</a>'
    return href


def main():
    st.set_page_config(
        page_title="供销云仓账单数据提取工具",
        page_icon="📊",
        layout="wide"
    )

    st.markdown("# 供销云仓账单数据提取工具")
    st.markdown("---")

    # 侧边栏 - 用于上传文件和显示操作状态
    with st.sidebar:
        st.header("操作面板")
        
        uploaded_files = st.file_uploader(
            "上传账单Excel文件",
            type=["xlsx", "xls"],
            accept_multiple_files=True
        )
        
        if st.button("清除结果", key="clear_button"):
            # 清除结果
            if "results" in st.session_state:
                st.session_state.results = []
                st.success("已清除所有结果！")
    
    # 主界面 - 显示结果表格
    if "results" not in st.session_state:
        st.session_state.results = []
    
    # 处理上传的文件
    if uploaded_files:
        processor = ExcelProcessor()
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_files = len(uploaded_files)
        processed_count = 0
        failed_files = []
        
        for i, file in enumerate(uploaded_files):
            try:
                # 更新进度
                progress = (i + 1) / total_files
                progress_bar.progress(progress)
                
                # 更新状态栏
                status_text.text(f"正在处理: {file.name} ({i + 1}/{total_files})")
                
                # 处理Excel文件
                file_bytes = file.read()
                result = processor.process_excel(file_bytes, file.name)
                
                # 检查结果是否已存在（按文件名）
                exists = False
                for r in st.session_state.results:
                    if r["文件名"] == result["文件名"]:
                        exists = True
                        break
                
                # 只有不存在时才添加
                if not exists:
                    st.session_state.results.append(result)
                
                processed_count += 1
                
            except Exception as e:
                failed_files.append((file.name, str(e)))
                st.error(f"处理文件 {file.name} 失败")
                print(f"处理文件 {file.name} 失败: {str(e)}")
        
        # 完成处理后的操作
        progress_bar.empty()
        
        if failed_files:
            status_text.text(f"已完成处理 {processed_count} 个文件，{len(failed_files)} 个文件失败")
            
            with st.expander("查看失败文件详情"):
                for i, (file_name, error) in enumerate(failed_files):
                    st.write(f"{i + 1}. {file_name}")
                    st.write(f"错误: {error.split('Traceback')[0]}")  # 只显示错误的第一部分
        else:
            status_text.text(f"已完成处理 {processed_count} 个文件")
    
    # 显示结果表格
    if st.session_state.results:
        st.markdown("## 处理结果")
        
        # 转换为DataFrame
        results_df = pd.DataFrame(st.session_state.results)
        
        # 显示表格
        st.dataframe(
            results_df,
            hide_index=True,
            use_container_width=True
        )
        
        # 下载按钮
        st.markdown(get_table_download_link(results_df), unsafe_allow_html=True)
    else:
        st.info("请上传账单Excel文件以开始处理")
    
    # 显示使用说明
    with st.expander("查看使用说明"):
        st.markdown("""
        ### 使用说明
        
        1. 在左侧操作面板点击"上传账单Excel文件"按钮上传一个或多个账单文件。
        2. 系统会自动处理上传的文件并提取关键数据。
        3. 处理结果将显示在表格中，包含以下字段：
           - 文件名
           - 月结账号
           - 账单周期
           - 当月单量
           - 费用(元)
           - 折扣/促销
           - 应付金额
           - 理赔费用合计
           - 账单总览金额
           - 特殊单票折扣
        4. 点击"下载Excel文件"链接可以将结果下载为Excel文件。
        5. 使用"清除结果"按钮可以清空当前结果。
        
        ### 注意事项
        
        - 支持的文件格式：.xlsx, .xls
        - 如果某些字段没有被正确提取，可能是因为文件结构与预期不符
        - 所有处理都在浏览器中完成，数据不会被上传到服务器
        """)
    
    # 页脚
    st.markdown("---")
    st.markdown("供销云仓账单数据提取工具 © 2025")

if __name__ == "__main__":
    main()
